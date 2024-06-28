#!/usr/bin/env python3 
import sys
import requests
import json
import datetime
import numpy as np
import os
import glob
import pandas as pd
import get_ovp
from dateutil.relativedelta import relativedelta
import pytz
import matplotlib.pyplot as plt
import seaborn as sns
from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

REPORTPATH = '/home/azureuser/HEDNOKPIs/daily_OVP/reports/'


# Define a function to generate Comments
def generate_comments(group):
    comments = []
    
    for phase, phase_group in group.groupby('phase'):
        num_events = phase_group['Number of events'].max()
        mean_kVA = phase_group['kVA'].mean()
        count_kVA = phase_group['kVA'].count()
        comment = f'Η φάση {phase} εμφάνισε {num_events} περιστατικά συνολικής διάρκειας {count_kVA} λεπτών με μέσο overpower {mean_kVA:.1f}%. '
        comments.append(comment)
    
    return ', '.join(comments)
    

# Define a custom aggregation function
def custom_agg(group):
    kVA_desc = group['kVA'].describe()[['min','mean','max','count']].to_frame().T
    kVA_desc.columns = [f'kVA_{col}' for col in kVA_desc.columns]
    
    max_events = group['Number of events'].max()
    
    result = pd.DataFrame({
        'Number of events': [max_events],
        'Comments': [generate_comments(group)]
    }, index=kVA_desc.index)
    combined_result = pd.concat([kVA_desc, result], axis=1)
    combined_result['phase'] = group['phase'].iloc[0]
    
    #return pd.concat([kVA_desc, result], axis=1)
    return combined_result




def read_excel_files(file_path):
    # Load the workbook and sheet
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # Read the data into a pandas dataframe
    df = pd.read_excel(file_path, engine='openpyxl')
    
    return df


def merge_old_xls(filename):
    new_filename = 'Historical_'+filename
    # Read the first and second Excel files with styles
    df1 = read_excel_files(REPORTPATH+filename)
    df2 = read_excel_files(REPORTPATH+'lastfile.xlsx')
    os.remove(REPORTPATH+'lastfile.xlsx')
   
    # Concatenate the dataframes with an empty row in between
    df_combined = pd.concat([df1, df2], ignore_index=True)
    df_combined = df_combined.dropna(how='all')
    
    
    writer = ExcelWriter(REPORTPATH+new_filename, engine='openpyxl')
    df_combined.to_excel(writer, index=False, sheet_name='Sheet1')
    # write to lastfile.xlsx
    df_combined.to_excel(REPORTPATH+'lastfile.xlsx', index=False, sheet_name='Sheet1',engine='openpyxl')
    
    merge_transf(df_combined, writer,1)
    
    
    
    #os.remove(REPORTPATH+'lastfile.xlsx')
    # create a copy of today's excel file
    #wb = load_workbook(REPORTPATH+new_filename)
    # Save it with a new name
    #wb.save(REPORTPATH+'lastfile.xlsx')
    
    return REPORTPATH+new_filename




def merge_transf(final_df, writer,mergeCells):

    # Get the workbook and active sheet
    wb = writer.book
    ws = wb['Sheet1']
    
    # Define bold font style
    bold_font = Font(bold=True)

    if mergeCells==1:
        # Find the columns for merging
        transformer_col_idx = final_df.columns.get_loc('Transformer') + 1
        
        # Apply bold font to 'Transformer' column
        for row in ws.iter_rows(min_row=2, max_row=len(final_df)+1, min_col=transformer_col_idx, max_col=transformer_col_idx):
            for cell in row:
                cell.font = bold_font
    
    
    phase_col_idx = final_df.columns.get_loc('phase') + 1
    #comments_col_idx = final_df.columns.get_loc('Comments') + 1
    
    # Track the current row number
    current_row = 2
    # Iterate through the DataFrame to merge cells for each transformer
    current_transformer = None
    start_row = None
    
    if mergeCells==1:
        for i, row in final_df.iterrows():
            row_num = current_row  # Current row in the Excel sheet
            transformer = row['Transformer']
            
            if transformer != current_transformer:
                if current_transformer is not None:
                    # Insert a blank row after each transformer (triplet of phases)
                    ws.insert_rows(current_row)
                    current_row += 1  # Move to the next row before merging
                    
                # Update current transformer and start row
                current_transformer = transformer
                start_row = current_row
            
            # Write the row to Excel
            for col_num, value in enumerate(row):
                ws.cell(row=current_row, column=col_num+1, value=value)
            
            current_row += 1
        
        
            # Merge cells for the current transformer and comments every triplet
            if (i + 1) % 3 == 0:
                ws.merge_cells(start_row=start_row, start_column=transformer_col_idx, end_row=current_row-1, end_column=transformer_col_idx)
                start_row = current_row  # Update start_row for the next triplet
    

    
    # Autofit column widths for all columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2  # Adding extra space

    # Save the workbook
    writer.save()





def process_info(df,phasedict,day,month,year):
    
    df = df.drop('% of time', axis=1)
    # Filter transformers first
    # Calculate sum of events per transformer
    sum_events = df.groupby('Transformer')['Number of events'].sum()
    
    # Filter transformers with sum of events > 0
    transformers_to_keep = sum_events[sum_events > 0].index
    
    # Filter original DataFrame based on transformers_to_keep
    df = df[df['Transformer'].isin(transformers_to_keep)]
    
    
    
    # Group by Transformer and then by phase
    all_results = []
    for transformer, group in df.groupby('Transformer'):
        phase_grouped = group.groupby('phase').apply(custom_agg).reset_index(drop=True)
        phase_grouped['Transformer'] = transformer
        all_results.append(phase_grouped)
    
    # Combine all results into a single DataFrame
    final_df = pd.concat(all_results).reset_index(drop=True)
    final_df.loc[final_df['Number of events']==0,['kVA_min','kVA_mean','kVA_max','kVA_count','Comments']]=''
    final_df['Date'] = year+'-'+month+'-'+day
    
    
    # Reorder columns to have 'Transformer' and 'phase' as the first columns
    cols = ['Transformer', 'phase'] + [col for col in final_df.columns if col not in ['Transformer', 'phase']]
    final_df = final_df[cols]
    final_df.rename(columns={'kVA_min':'Minimum overpower %','kVA_mean':'Average overpower %','kVA_max':'Maximum overpower %','kVA_count':'Total minutes'},inplace=True)
    # Write to Excel
    if (day=='01') and (month=='01'):
        filename = 'lastfile.xlsx'
    else:
        filename = 'Overpower_analysis_'+str(year)+'_'+str(month)+'_'+str(day)+'.xlsx'
    writer = ExcelWriter(REPORTPATH+filename, engine='openpyxl')
    final_df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    
    merge_transf(final_df,writer,0)
    
    return filename
 
 
 
def ultimatefig(df, day, month, year,monthdayfig,phasedict):
    
    fig, axes = plt.subplots(2, 1, figsize=(14, 8), gridspec_kw={'height_ratios': [3, 1]})

    # First subplot: boxplot
    sns.boxplot(ax=axes[0], data=df, x='Transformer', y='kVA',hue='phase' ,palette='Set2')
    axes[0].set_xlabel('')
    axes[0].set_ylabel('Overpower kVA %')
    axes[0].set_title('Overpower boxplots '+day+'/'+month+'/'+year)
    axes[0].tick_params(axis='x', rotation=45)
    #axes[0].legend(title='')
    #axes[0].legend().remove()
    
    # Second subplot: another plot (e.g., bar plot)
    sns.barplot(ax=axes[1], data=df, x='Transformer', y='% of time',hue='phase', palette='Set2')
    axes[1].set_xlabel('')
    axes[1].set_ylabel('% of time')
    axes[1].set_title('Time % of overpower alarms')
    axes[1].tick_params(axis='x', rotation=45)
    # axes[1].legend(title='Month')
    #axes[1].legend().remove()

    # Adjust layout
    plt.tight_layout()
    
    figname = monthdayfig+'ovp_boxplot_'+year+'_'+month+'_'+day+'.png'
    plt.savefig(figname, dpi=300)
    
    return figname

 
      
      
def recursion(start_time, end_time, day, month, year):

    mainpath = '/home/azureuser/HEDNOKPIs/daily_OVP/'
    monthdayfig = mainpath+'figures/'+month+'/'
    if not os.path.exists(monthdayfig):
        os.makedirs(monthdayfig)
    
    
    monthdaypath = mainpath+'processed_data/'+month+'/'+day+'/'
    if not os.path.exists(monthdaypath):
        os.makedirs(monthdaypath)
    os.chdir(monthdaypath)
    
    
    
    address = 'http://localhost:8080'
    r = requests.post(address + "/api/auth/login",
                      json={'username': 'meazonpro@meazon.com', 'password': 'meazonpro1'}).json()
    
    acc_token = 'Bearer' + ' ' + r['token']
    
    entityId = '47545f30-5b7f-11ee-b2c9-653b42f73605' # DEDDHE ATHINAS
    r1 = requests.get(url=address + "/api/entityGroup/"+entityId+"/entities?pageSize=1000&page=0",headers={'Content-Type': 'application/json', 
'Accept': '*/*', 'X-Authorization': acc_token}).json()
    
    alltransf = []
    
    for i in range(0,len(r1['data'])):
     #   os.chdir('/home/azureuser/deddhePDF/')
        assetid = r1['data'][i]['id']['id']
        assetname = r1['data'][i]['name']
        print(assetname)
    
        if assetname[0]!='0':
        
            r2 = requests.get(url=address + "/api/relations/info?fromId="+assetid+"&fromType=ASSET",headers={'Content-Type': 'application/json', 'Accept': '*/*', 'X-Authorization': acc_token}).json()
            
            
            for j in range(0, len(r2)):
                device = r2[j]['toName']
                
                if device[:3]=='102':
                    
                    #print(device)
                    
                    #try:
                    [devid, acc_token, label] = get_ovp.get_dev_info(device, address)
                    
                    alltransf.append(label)
                    #get_ovp.main(device, start_time, end_time, assetname,devid, acc_token, label, mainpath)
    set1 = set(alltransf)
    
    
    # Create figures
    files = os.listdir(monthdaypath)
    phasedict = {'L1':'A','L2':'B','L3':'C'}
    pwrdf = pd.DataFrame([])
    
    mydict = {}
    for ph in ['L1','L2','L3']:
        df = pd.DataFrame([])
        matching_files = [f for f in files if os.path.basename(f).upper().endswith(phasedict[ph]+'.XLSX')]
        for filename in matching_files:
            tmp = pd.read_excel(monthdaypath+filename, engine='openpyxl')
            tmp = tmp.drop('apwr'+phasedict[ph], axis=1)
            tmp = tmp.rename(columns={'exceed_perc'+phasedict[ph]:'kVA','% of time'+phasedict[ph]:'% of time','events'+phasedict[ph]:'Number of events'})
            tmp['kVA'] = pd.to_numeric(tmp['kVA'])
            
            df = pd.concat([df, tmp])
        
        if not df.empty:  
            set2 = set(list(df['Transformer'].unique()))
            missing_transf = list(set1.symmetric_difference(set2))
        else:
            missing_transf = list(set1)
        
        common_df = pd.DataFrame([t for t in missing_transf], columns=['Transformer'])
        common_df['% of time'] = 0.0
        common_df['kVA'] = 80#np.nan
        common_df['Number of events'] = 0
        df = pd.concat([df, common_df])
        df['phase'] = ph
        pwrdf = pd.concat([pwrdf,df])
        
            
    
    if not pwrdf.empty:
        pwrdf['customT'] = pd.Categorical(pwrdf['Transformer'], categories=alltransf, ordered=True)
        pwrdf['customP'] = pd.Categorical(pwrdf['phase'], categories=['L1','L2','L3'], ordered=True)
        pwrdf = pwrdf.sort_values(by=['customT','customP'])
        pwrdf = pwrdf.drop(['customT','customP'],axis=1)
        pwrdf = pwrdf.reset_index(drop=True)
        
        if pwrdf['Number of events'].sum()>0:
            figname = ultimatefig(pwrdf, day, month, year,monthdayfig,phasedict)
            #filename = process_info(pwrdf,phasedict,day,month,year) 
            #if (day=='01') and (month=='01'):
            #    print('First day!')
            #else:
            #    dataname = merge_old_xls(filename)
                
                 

def main():

        # Define the timezone
    athens_tz = pytz.timezone('Europe/Athens')
    
    # Define the start and end dates
    start_date = datetime.datetime(2024, 1, 1)
    end_date = datetime.datetime(2024, 6, 17)
    
    # Iterate over each day in the date range
    current_date = start_date
    while current_date <= end_date:
        # Get the start of the day in Athens timezone
        start_of_day = athens_tz.localize(datetime.datetime(current_date.year, current_date.month, current_date.day, 0, 0, 0))
        # Get the end of the day in Athens timezone (23:59:59 of the same day)
        end_of_day = athens_tz.localize(datetime.datetime(current_date.year, current_date.month, current_date.day) + datetime.timedelta(days=1))
    
        # Convert to Unix timestamps in milliseconds
        start_time = int(start_of_day.timestamp() * 1000)
        end_time = int(end_of_day.timestamp() * 1000)
        day = str(current_date.day)
        if len(day)==1:
            day = '0'+day
        month = str(current_date.month)
        if len(month)==1:
            month = '0'+month
        year = str(current_date.year)
        print(f"Date: {current_date.strftime('%Y-%m-%d')}")
        
        
    
        # Move to the next day
        current_date += datetime.timedelta(days=1)
        recursion(start_time, end_time, day, month, year)
if __name__ == '__main__':
    sys.exit(main())