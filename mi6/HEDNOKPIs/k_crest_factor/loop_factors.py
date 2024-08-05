#!/usr/bin/env python3 
import sys
import requests
import json
import datetime
import numpy as np
import os
import glob
import pandas as pd
import get_raw
from dateutil.relativedelta import relativedelta
import pytz
import matplotlib.pyplot as plt
import seaborn as sns
from pandas import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows


REPORTPATH = '/home/azureuser/HEDNOKPIs/k_crest_factor/reports/'

    

# Define a custom aggregation function
def custom_agg(group):
    kVA_desc = group['kVA'].describe()[['min','mean','max','count']].to_frame().T
    kVA_desc.columns = [f'kVA_{col}' for col in kVA_desc.columns]
    
    max_events = group['Number of events'].max()
    
    result = pd.DataFrame({
        'Number of events': [max_events],
        'Comments': [generate_comments(group)]
    }, index=kVA_desc.index)
    combined_result = pd.concat([kVA_desc, result], axis=1)
    combined_result['phase'] = group['phase'].iloc[0]
    
    #return pd.concat([kVA_desc, result], axis=1)
    return combined_result




def merge_transf(final_df, writer,mergeCells):

    # Get the workbook and active sheet
    wb = writer.book
    ws = wb['Sheet1']
    
    # Define bold font style
    bold_font = Font(bold=True)

    if mergeCells==1:
        # Find the columns for merging
        transformer_col_idx = final_df.columns.get_loc('Transformer') + 1
        
        # Apply bold font to 'Transformer' column
        for row in ws.iter_rows(min_row=2, max_row=len(final_df)+1, min_col=transformer_col_idx, max_col=transformer_col_idx):
            for cell in row:
                cell.font = bold_font
    
    
    phase_col_idx = final_df.columns.get_loc('phase') + 1
    #comments_col_idx = final_df.columns.get_loc('Comments') + 1
    
    # Track the current row number
    current_row = 2
    # Iterate through the DataFrame to merge cells for each transformer
    current_transformer = None
    current_date = None
    start_row = None
    
    if mergeCells==1:
        for i, row in final_df.iterrows():
            row_num = current_row  # Current row in the Excel sheet
            transformer = row['Transformer']
            rowdate = row['Date']
            
            if (transformer != current_transformer) or (rowdate != current_date):
                if current_transformer is not None:
                    # Insert a blank row after each transformer (triplet of phases)
                    ws.insert_rows(current_row)
                    current_row += 1  # Move to the next row before merging
                    
                # Update current transformer and start row
                current_transformer = transformer
                current_date = rowdate
                start_row = current_row
            
            # Write the row to Excel
            for col_num, value in enumerate(row):
                ws.cell(row=current_row, column=col_num+1, value=value)
            
            current_row += 1
        
        
            # Merge cells for the current transformer and comments every triplet
            if (i + 1) % 3 == 0:
                ws.merge_cells(start_row=start_row, start_column=transformer_col_idx, end_row=current_row-1, end_column=transformer_col_idx)
                start_row = current_row  # Update start_row for the next triplet
    

    
    # Autofit column widths for all columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2  # Adding extra space

    # Save the workbook
    writer.save()



def process_info(df,phasedict,day,month,year):
    
    df = df.drop('% of time', axis=1)
    # Filter transformers first
    # Calculate sum of events per transformer
    sum_events = df.groupby('Transformer')['Number of events'].sum()
    
    # Filter transformers with sum of events > 0
    transformers_to_keep = sum_events[sum_events > 0].index
    
    # Filter original DataFrame based on transformers_to_keep
    df = df[df['Transformer'].isin(transformers_to_keep)]
    
    
    
    # Group by Transformer and then by phase
    all_results = []
    for transformer, group in df.groupby('Transformer'):
        phase_grouped = group.groupby('phase').apply(custom_agg).reset_index(drop=True)
        phase_grouped['Transformer'] = transformer
        all_results.append(phase_grouped)
    
    # Combine all results into a single DataFrame
    final_df = pd.concat(all_results).reset_index(drop=True)
    final_df.loc[final_df['Number of events']==0,['kVA_min','kVA_mean','kVA_max','kVA_count','Comments']]=''
    final_df['Date'] = year+'-'+month+'-'+day
    
    
    # Reorder columns to have 'Transformer' and 'phase' as the first columns
    cols = ['Transformer', 'phase'] + [col for col in final_df.columns if col not in ['Transformer', 'phase']]
    final_df = final_df[cols]
    final_df.rename(columns={'kVA_min':'Minimum overpower %','kVA_mean':'Average overpower %','kVA_max':'Maximum overpower %','kVA_count':'Total minutes'},inplace=True)
    # Write to Excel
    filename = 'Overpower_analysis_'+str(year)+'_'+str(month)+'_'+str(day)+'.xlsx'
    writer = ExcelWriter(REPORTPATH+filename, engine='openpyxl')
    final_df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    
    merge_transf(final_df,writer,0)
    
    return filename
  
def ultimatefig_kfac(df, day, month, year,monthdayfig):
    fig, ax = plt.subplots(figsize=(14, 8))

    # First subplot: boxplot
    sns.boxplot(ax=ax, data=df, x='Transformer', y='k-factor',palette='Set2')
    ax.set_xlabel('')
    ax.set_ylabel('K factor')
    ax.set_title('K factor boxplots (upon condition: V*I>0.8*nominal) '+year+'/'+month+'/'+day)
    ax.tick_params(axis='x', rotation=45)
    #axes[0].legend(title='')
    #axes[0].legend().remove()
    
    
    # Adjust layout
    plt.tight_layout()
    
    figname = monthdayfig+'k_fac_boxplot_'+year+'_'+month+'_'+day+'.png'
    plt.savefig(figname, dpi=300)
    
    return figname
    
    
def ultimatefig_curNd(df, day, month, year,monthdayfig):
    fig, ax = plt.subplots(figsize=(14, 8))

    # First subplot: boxplot
    sns.boxplot(ax=ax, data=df, x='Transformer', y='curN_perc',palette='Set2')
    ax.set_xlabel('')
    ax.set_ylabel('Summed curN to Sensed curN %')
    ax.set_title('Neutral current (curNsu/curNse) boxplots '+month+'/'+year)
    ax.tick_params(axis='x', rotation=45)
    #axes[0].legend(title='')
    #axes[0].legend().remove()
    
    
    # Adjust layout
    plt.tight_layout()
    
    figname = monthdayfig+'curN_boxplot_summed_to_sensed_'+year+'_'+month+'.png'
    plt.savefig(figname, dpi=300)
    return figname

def ultimatefig(df, day, month, year,monthdayfig,phasedict):
    
    # fig, axes = plt.subplots(1, 1, figsize=(14, 8), gridspec_kw={'height_ratios': [3, 1]})
    fig, ax = plt.subplots(figsize=(14, 8))

    # First subplot: boxplot
    sns.boxplot(ax=ax, data=df, x='Transformer', y='Crest factor',hue='phase' ,palette='Set2')
    ax.set_xlabel('')
    ax.set_ylabel('Crest factor')
    ax.set_title('Crest factor boxplots (upon condition: V*I>0.8*nominal) '+month+'/'+year)
    ax.tick_params(axis='x', rotation=45)
    #axes[0].legend(title='')
    #axes[0].legend().remove()
    
    
    # Adjust layout
    plt.tight_layout()
    
    figname = monthdayfig+'crest_fac_boxplot_'+year+'_'+month+'_'+day+'.png'
    plt.savefig(figname, dpi=300)
    
    return figname
    



def main(arguments):
    #var = 'kfac'
    var = arguments[1]
    print(var)
    mainpath = '/home/azureuser/HEDNOKPIs/k_crest_factor/'
    #os.chdir(mainpath)
    end_time = datetime.datetime.now(pytz.timezone('Europe/Athens'))
    
    end_time = end_time - datetime.timedelta(hours=end_time.hour, minutes=end_time.minute, seconds=end_time.second,
                                                 microseconds=end_time.microsecond)
    
    start_time = end_time +relativedelta(days=-1)
    day = str(start_time.day)
    if len(day)==1:
        day = '0'+day
    month = str(start_time.month)
    if len(month)==1:
        month = '0'+month
    year = str(start_time.year)
    
    start_time = str(int(start_time.timestamp()) * 1000)
    end_time = str(int(end_time.timestamp()) * 1000)
    print(start_time,end_time)
    
    #day = '01'
    #month = '07'
    #year = '2024'
    #start_time = '1719781200000' # July 1st
    #end_time = '1721682000000' # July 23
    #start_time = '1719262800000'
    #end_time = '1721138338000'
        
    
    monthdayfig = mainpath+'figures/'+month+'/'
    if not os.path.exists(monthdayfig):
        os.makedirs(monthdayfig)
        
    monthdaypath = mainpath+'processed_data_'+var+'/'+month+'/'+day+'/'
    if not os.path.exists(monthdaypath):
        os.makedirs(monthdaypath)
    os.chdir(monthdaypath)
    
    
    address = 'http://localhost:8080'
    r = requests.post(address + "/api/auth/login",
                      json={'username': 'meazonpro@meazon.com', 'password': 'meazonpro1'}).json()
    
    acc_token = 'Bearer' + ' ' + r['token']
    
    entityId = '47545f30-5b7f-11ee-b2c9-653b42f73605' # DEDDHE ATHINAS
    r1 = requests.get(url=address + "/api/entityGroup/"+entityId+"/entities?pageSize=1000&page=0",headers={'Content-Type': 'application/json', 
'Accept': '*/*', 'X-Authorization': acc_token}).json()
    
    alltransf = []
    
    for i in range(0,len(r1['data'])):
     #   os.chdir('/home/azureuser/deddhePDF/')
        assetid = r1['data'][i]['id']['id']
        assetname = r1['data'][i]['name']
        print(assetname)
    
        if assetname[0]!='0':
        
            r2 = requests.get(url=address + "/api/relations/info?fromId="+assetid+"&fromType=ASSET",headers={'Content-Type': 'application/json', 'Accept': '*/*', 'X-Authorization': acc_token}).json()
            
            
            for j in range(0, len(r2)):
                device = r2[j]['toName']
                
                if device[:3]=='102':
                    
                    #print(device)
                    
                    #try:
                    [devid, acc_token, label] = get_raw.get_dev_info(device, address)
                    
                    alltransf.append(label)
                
                    get_raw.main(device, start_time, end_time, assetname,devid, acc_token, label,var)
    set1 = set(alltransf)
    
    
    # Create figures
    files = os.listdir(monthdaypath)
    phasedict = {'L1':'A','L2':'B','L3':'C'}
    
    
    if var=='crest':
        crestdf = pd.DataFrame([])
        
        for ph in ['L1','L2','L3']:
            df = pd.DataFrame([])
            matching_files = [f for f in files if os.path.basename(f).upper().endswith(phasedict[ph]+'.XLSX')]
            for filename in matching_files:
                tmp = pd.read_excel(monthdaypath+filename, engine='openpyxl')
                tmp = tmp.drop('apwr'+phasedict[ph], axis=1)
                tmp = tmp.rename(columns={'scre'+phasedict[ph]:'Crest factor'})
                tmp['Crest factor'] = pd.to_numeric(tmp['Crest factor'])
                
                df = pd.concat([df, tmp])
            
            if not df.empty:  
                set2 = set(list(df['Transformer'].unique()))
                missing_transf = list(set1.symmetric_difference(set2))
            else:
                missing_transf = list(set1)
            
            common_df = pd.DataFrame([t for t in missing_transf], columns=['Transformer'])
            common_df['Crest factor'] = 1.0#np.nan
            df = pd.concat([df, common_df])
            df['phase'] = ph
            crestdf = pd.concat([crestdf,df])
            
            
        
        if not crestdf.empty:
            crestdf['customT'] = pd.Categorical(crestdf['Transformer'], categories=alltransf, ordered=True)
            crestdf['customP'] = pd.Categorical(crestdf['phase'], categories=['L1','L2','L3'], ordered=True)
            crestdf = crestdf.sort_values(by=['customT','customP'])
            crestdf = crestdf.drop(['customT','customP'],axis=1)
            crestdf = crestdf.reset_index(drop=True)
            
            # if pwrdf['Number of events'].sum()>0:
            figname = ultimatefig(crestdf, day, month, year,monthdayfig,phasedict)         

    if var=='kfac':
        kfacdf = pd.DataFrame([])
        df = pd.DataFrame([])
        matching_files = [f for f in files if os.path.basename(f).upper().endswith('.XLSX')]
        for filename in matching_files:
            tmp = pd.read_excel(monthdaypath+filename, engine='openpyxl')
            #tmp = tmp.drop('apwr', axis=1)
            tmp = tmp.rename(columns={'kfac':'k-factor'})
            tmp['k-factor'] = pd.to_numeric(tmp['k-factor'])
            
            df = pd.concat([df, tmp])
        
        if not df.empty:  
            set2 = set(list(df['Transformer'].unique()))
            missing_transf = list(set1.symmetric_difference(set2))
        else:
            missing_transf = list(set1)
        
        common_df = pd.DataFrame([t for t in missing_transf], columns=['Transformer'])
        common_df['k-factor'] = 1.0#np.nan
        df = pd.concat([df, common_df])
        kfacdf = pd.concat([kfacdf,df])
        if not kfacdf.empty:
            kfacdf['customT'] = pd.Categorical(kfacdf['Transformer'], categories=alltransf, ordered=True)
            
            kfacdf = kfacdf.sort_values(by='customT')
            kfacdf = kfacdf.drop('customT',axis=1)
            kfacdf = kfacdf.reset_index(drop=True)
            
            if kfacdf['k-factor'].max()>1.1:
                ultimatefig_kfac(kfacdf, day, month, year,monthdayfig)
            
    if var=='curNd':
        curdf = pd.DataFrame([])
        df = pd.DataFrame([])
        matching_files = [f for f in files if os.path.basename(f).upper().endswith('.XLSX')]
        for filename in matching_files:
            tmp = pd.read_excel(monthdaypath+filename, engine='openpyxl')
            
            tmp['curN_perc'] = 100*(tmp['curNsu']/tmp['curNse'])
            tmp = tmp.loc[tmp['curN_perc']<500]
            #tmp['curN_perc'] = pd.to_numeric(tmp['curN_perc'])
            #tmp = tmp.loc[(tmp['curNd']<100) & (tmp['curNd']>-100)]
            tmp = tmp[['curN_perc','Transformer']]
            
            df = pd.concat([df, tmp])
        
        if not df.empty:  
            set2 = set(list(df['Transformer'].unique()))
            missing_transf = list(set1.symmetric_difference(set2))
        else:
            missing_transf = list(set1)
        
        common_df = pd.DataFrame([t for t in missing_transf], columns=['Transformer'])
        common_df['curN_perc'] = 0.0#np.nan
        df = pd.concat([df, common_df])
        curdf = pd.concat([curdf,df])
        if not curdf.empty:
            curdf['customT'] = pd.Categorical(curdf['Transformer'], categories=alltransf, ordered=True)
            
            curdf = curdf.sort_values(by='customT')
            curdf = curdf.drop('customT',axis=1)
            curdf = curdf.reset_index(drop=True)
            
            # if pwrdf['Number of events'].sum()>0:
            ultimatefig_curNd(curdf, day, month, year,monthdayfig)      

if __name__ == '__main__':
    sys.exit(main(sys.argv))